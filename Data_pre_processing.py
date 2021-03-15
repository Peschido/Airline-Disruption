from openpyxl import load_workbook 
import numpy as np 

dataset = "Datasets/Programming_Set/"

#####   AIRCRAFT FILE
wb = load_workbook(dataset+"aircraft.xlsx", data_only=True) 
ws = wb.active 

aircraft_raw = np.array([[i.value for i in j] for j in ws ['A1':'A255']]) 
aircraft = np.array([["                                      " \
           for i in range(len(aircraft_raw[254][0].split(" ")))] for j in range(len(aircraft_raw))])
for i in range(len(aircraft_raw)): 
    for j in range(len(aircraft_raw[i][0].split(" "))):
        aircraft[i][j] = aircraft_raw[i][0].split(" ")[j]
    
#####   AIRPORTS FILE
wb = load_workbook(dataset+"airports.xlsx", data_only=True) 
ws = wb.active 

airports_raw = np.array([[i.value for i in j] for j in ws ['A1':'A44']]) 
airports = np.array([["                                      " \
           for i in range(len(airports_raw[23][0].split(" ")) -1)] for j in range(len(airports_raw))])
for i in range(len(airports_raw)): 
    for j in range(len(airports_raw[i][0].split(" "))-1):
        airports[i][j] = airports_raw[i][0].split(" ")[j]
        
#####   CONFIGURATION FILE
wb = load_workbook(dataset+"config.xlsx", data_only=True) 
ws = wb.active 

config_raw = np.array([[i.value for i in j] for j in ws ['A1':'A7']]) 
config = np.array([["                                      " \
           for i in range(len(config_raw[4][0].split(" ")) -1)] for j in range(len(config_raw))])
for i in range(len(config_raw)): 
    for j in range(len(config_raw[i][0].split(" "))-1):
        config[i][j] = config_raw[i][0].split(" ")[j]
        
#####   DISTANCE FILE
wb = load_workbook(dataset+"dist.xlsx", data_only=True) 
ws = wb.active 

dist_raw = np.array([[i.value for i in j] for j in ws ['A1':'A1892']]) 
dist = np.array([["                                      " \
           for i in range(len(dist_raw[1][0].split(" ")) -1)] for j in range(len(dist_raw))])
for i in range(len(dist_raw)): 
    for j in range(len(dist_raw[i][0].split(" "))-1):
        dist[i][j] = dist_raw[i][0].split(" ")[j]
    
        
#####   FLIGHTS FILE
wb = load_workbook(dataset+"flights.xlsx", data_only=True) 
ws = wb.active 

flights_raw = np.array([[i.value for i in j] for j in ws ['A1':'A1422']]) 
flights = np.array([["                                      " \
           for i in range(len(flights_raw[1][0].split(" ")) -1)] for j in range(len(flights_raw))])
for i in range(len(flights_raw)): 
    for j in range(len(flights_raw[i][0].split(" "))-1):
        flights[i][j] = flights_raw[i][0].split(" ")[j]
        
#####   ITINERARIES FILE
wb = load_workbook(dataset+"itineraries.xlsx", data_only=True) 
ws = wb.active 

itineraries_raw = np.array([[i.value for i in j] for j in ws ['A1':'A11213']]) 
itineraries = np.array([["                                      " \
           for i in range(len(itineraries_raw[1][0].split(" ")) -1)] for j in range(len(itineraries_raw))])
for i in range(len(itineraries_raw)): 
    for j in range(len(itineraries_raw[i][0].split(" "))-1):
        itineraries[i][j] = itineraries_raw[i][0].split(" ")[j]
        
#####   POSITIONS FILE
wb = load_workbook(dataset+"position.xlsx", data_only=True) 
ws = wb.active 

position_raw = np.array([[i.value for i in j] for j in ws ['A1':'A44']]) 
position = np.array([["                                      " \
           for i in range(len(position_raw[8][0].split(" ")) -1)] for j in range(len(position_raw))])
for i in range(len(position_raw)): 
    for j in range(len(position_raw[i][0].split(" "))-1):
        position[i][j] = position_raw[i][0].split(" ")[j]
        
#####   ROTATIONS FILE
wb = load_workbook(dataset+"rotations.xlsx", data_only=True) 
ws = wb.active 

rotations_raw = np.array([[i.value for i in j] for j in ws ['A1':'A2844']]) 
rotations = np.array([["                                      " \
           for i in range(len(rotations_raw[8][0].split(" ")) -1)] for j in range(len(rotations_raw))])
for i in range(len(rotations_raw)): 
    for j in range(len(rotations_raw[i][0].split(" "))-1):
        rotations[i][j] = rotations_raw[i][0].split(" ")[j]